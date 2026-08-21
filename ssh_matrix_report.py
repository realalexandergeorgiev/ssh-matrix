#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ssh_matrix_report.py - baut aus detail.csv (Output von ssh_matrix.py) eine
Matrix-CSV und eine Excel-Datei (report.xlsx) mit Dropdown-Suche VON -> NACH.

Sheets in report.xlsx:
  Suche    - Dropdowns fuer VON/NACH (Datenvalidierung, Vorschlaege beim Tippen),
             Ergebnis via INDEX/MATCH auf dem Matrix-Sheet, Legende.
  Matrix   - Pivot: Zeilen = Quelle, Spalten = Ziel, Zelle = Status-Code mit Farbe.
  Detail   - alle Testzeilen mit AutoFilter und bedingter Formatierung.
  Subnetze - Statistik pro /24-Gruppe.
  Quelle   - Endpunkt-Mapping (IP, Port, Label, /24, /16).
  Liste    - Referenzliste der Endpunkte (Datenquelle fuer die Dropdowns).
"""

import argparse
import csv
import ipaddress
import logging
import os
import sys
from collections import Counter, defaultdict, deque

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("FEHLER: openpyxl fehlt. Auf dem Kali-Host installieren:", file=sys.stderr)
    print("  sudo apt install -y python3-openpyxl", file=sys.stderr)
    sys.exit(2)

DEFAULT_PORT = 22

VERSION = "v2.0.6"
AUTHOR = "Alex & DeepSeek"


def print_banner(stream=sys.stderr) -> None:
    """Start-Banner mit Version und Autoren-Hinweis."""
    line = "=" * 44
    print(line, file=stream)
    print(f"   SSH-Matrix-Tester {VERSION} - Report-Generator", file=stream)
    print(f"   entwickelt von {AUTHOR}", file=stream)
    print(line, file=stream)


CODE = {
    "auth_ok": "OK",
    "auth_fail": "AUTH",
    "port_open": "PORT",
    "port_closed": "CLOSED",
    "net_unreachable": "UNREACH",
    "source_unreachable": "SRCERR",
    "no_tool": "NOTOOL",
    "tool_error": "ERR",
    "unclear": "?",
    "skipped": "SKIP",
}

# Status -> (Hintergrund, Schrift)
STYLE = {
    "OK":     ("C6EFCE", "006100"),
    "AUTH":   ("FFEB9C", "9C6500"),
    "PORT":   ("DDEBF7", "1F4E79"),
    "CLOSED": ("D9D9D9", "404040"),
    "UNREACH": ("FFC7CE", "9C0006"),
    "SRCERR": ("F8CBAD", "833C00"),
    "NOTOOL": ("FCE4D6", "974706"),
    "ERR":    ("FCE4D6", "974706"),
    "?":      ("FFFFFF", "000000"),
    "SKIP":   ("E7E6E6", "595959"),
}


def ep_label(ip: str, port: str) -> str:
    return ip if int(port) == DEFAULT_PORT else f"{ip}:{port}"


def fmt_num(n: int) -> str:
    """Zahl mit Punkt-Tausendertrennung."""
    return f"{n:,}".replace(",", ".")


# Streaming-Architektur: detail.csv wird NIE komplett in den RAM geladen.
# Status je Paar landet in einem codes-Bytearray (1 Byte/Paar) statt in
# einem Riesen-Dict (~2,5 GB bei 12,6 Mio. Paaren).

STATUS_TO_CODE = {st: i + 1 for i, st in enumerate(CODE)}
CODE_TO_SHORT = {v: CODE[k] for k, v in STATUS_TO_CODE.items()}


def _register_ep(eps: dict, r: dict, side: str) -> dict:
    key = (r[f"{side}_ip"], int(r[f"{side}_port"]))
    e = eps.get(key)
    if e is None:
        e = {
            "ip": r[f"{side}_ip"],
            "port": int(r[f"{side}_port"]),
            "label": r.get(f"{side}_label", "") or "",
            "n24": r.get(f"{side}_24", r.get("src_24", "")),
            "n16": r.get(f"{side}_16", r.get("src_16", "")),
            "net": r.get(f"{'src' if side == 'source' else 'tgt'}_net",
                         r.get(f"{side}_24", "")),
        }
        eps[key] = e
    elif not e["label"] and r.get(f"{side}_label"):
        e["label"] = r[f"{side}_label"]
    return e


def analyze_detail(path: str) -> dict:
    """Pass 1 (streaming): Endpunkte, Netz-Aggregation, Status-Verteilung,
    Endpunkt-Statistiken. Return-Dict mit allen Daten."""
    eps = {}
    agg = defaultdict(lambda: {"ok": 0, "port_only": 0, "failed": 0,
                               "tested": 0, "skipped": 0})
    nets = set()
    counts = Counter()
    src_ok = set()          # (ip, port) mit auth_ok als Quelle
    tgt_reach = set()       # (ip, port) erreichbar als Ziel (auth_ok/port_open)
    tgt_auth = set()        # (ip, port) mit auth_ok als Ziel
    skipped_by_n24 = Counter()
    rows_total = 0
    with open(path, newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            rows_total += 1
            s = _register_ep(eps, r, "source")
            t = _register_ep(eps, r, "target")
            st = r["status"]
            counts[st] += 1
            s_key = (s["ip"], s["port"])
            t_key = (t["ip"], t["port"])
            if st == "auth_ok":
                src_ok.add(s_key)
                tgt_auth.add(t_key)
            if st in ("auth_ok", "port_open"):
                tgt_reach.add(t_key)
            if st == "skipped":
                skipped_by_n24[r.get("src_24", "") or r.get("src_net", "")] += 1
            s_net = r.get("src_net") or r.get("src_24", "")
            t_net = r.get("tgt_net") or r.get("tgt_24", "")
            if s_net and t_net:
                nets.add(s_net)
                nets.add(t_net)
                a = agg[(s_net, t_net)]
                if st == "skipped":
                    a["skipped"] += 1
                else:
                    a["tested"] += 1
                    if st == "auth_ok":
                        a["ok"] += 1
                    elif st == "port_open":
                        a["port_only"] += 1
                    else:
                        a["failed"] += 1
    eps_sorted = sorted(eps.values(),
                        key=lambda e: (ipaddress.IPv4Address(e["ip"]).packed,
                                       e["port"]))
    id_of = {(e["ip"], e["port"]): i for i, e in enumerate(eps_sorted)}
    sorted_nets = sorted(nets,
                         key=lambda n: ipaddress.IPv4Network(n, strict=False)
                         .network_address.packed)
    return {
        "eps": eps_sorted,
        "id_of": id_of,
        "agg": dict(agg),
        "nets": sorted_nets,
        "counts": counts,
        "src_ok": src_ok,
        "tgt_reach": tgt_reach,
        "tgt_auth": tgt_auth,
        "skipped_by_n24": skipped_by_n24,
        "rows_total": rows_total,
    }


def pass2_codes(path: str, id_of: dict, n: int, detail_max: int) -> tuple:
    """Pass 2 (streaming): fuellt das codes-Bytearray (alle Zeilen) und
    sammelt die ersten detail_max Zeilen fuer das Detail-Sheet.
    Return: (codes, detail_header, detail_rows, detail_truncated)."""
    codes = bytearray(n * n)
    detail_rows = []
    detail_header = None
    truncated = 0
    with open(path, newline="", encoding="utf-8") as fh:
        rd = csv.DictReader(fh)
        detail_header = rd.fieldnames
        for r in rd:
            try:
                s = id_of.get((r["source_ip"], int(r["source_port"])))
                t = id_of.get((r["target_ip"], int(r["target_port"])))
            except (KeyError, ValueError):
                continue
            if s is not None and t is not None and s != t:
                codes[s * n + t] = STATUS_TO_CODE.get(r["status"], 0)
            if len(detail_rows) < detail_max:
                detail_rows.append(r)
            else:
                truncated += 1
    return codes, detail_header, detail_rows, truncated


# ---------------------------------------------------------------- Pfadfindung
# Gerichtete Kanten: auth_ok (verifiziert, Prioritaet) und
# auth_fail/port_open ("reachable", sekundaer). Alle anderen Status = keine
# Kante. BFS mit Hop-Limit + Caps; Subnetz-Pfade als Fallback.

CODE_AUTH = STATUS_TO_CODE["auth_ok"]
CODE_REACH = {STATUS_TO_CODE[s] for s in ("auth_fail", "port_open")}
FRONTIER_CAP = 100000  # Besuchte Knoten pro Quelle -> dichte Graphen begrenzen


def build_adjacency(codes: bytearray, n: int) -> tuple:
    auth_adj = [[] for _ in range(n)]
    reach_adj = [[] for _ in range(n)]
    for i in range(n):
        base = i * n
        ai = auth_adj[i]
        ri = reach_adj[i]
        for j in range(n):
            if i == j:
                continue
            c = codes[base + j]
            if c == CODE_AUTH:
                ai.append(j)
                ri.append(j)
            elif c in CODE_REACH:
                ri.append(j)
    return auth_adj, reach_adj


def _bfs(adj, start: int, n: int, max_hops: int, frontier_cap: int):
    """BFS mit Hop-Limit. Return (prev, reason); reason 'frontier' wenn das
    Frontier-Cap erreicht wurde."""
    prev = [-1] * n
    visited = bytearray(n)
    visited[start] = 1
    queue = deque([start])
    visited_count = 1
    for _depth in range(max_hops):
        if not queue:
            break
        nxt = []
        for u in queue:
            for v in adj[u]:
                if not visited[v]:
                    visited[v] = 1
                    prev[v] = u
                    nxt.append(v)
                    visited_count += 1
        queue = nxt
        if visited_count > frontier_cap:
            return prev, "frontier"
    return prev, "ok"


def _path_to(prev: list, src: int, tgt: int):
    if prev[tgt] == -1:
        return None
    path = [tgt]
    cur = tgt
    while cur != src:
        cur = prev[cur]
        if cur == -1:
            return None
        path.append(cur)
    path.reverse()
    return path


def find_ip_paths(codes, eps, n, max_hops, max_paths, log) -> tuple:
    """BFS pro Quelle: erst auth_ok-Kanten (verifiziert), dann
    auth_ok+reachable. Nur Paare OHNE direkte Kante. Return (rows, capped)."""
    rows = []
    capped = False
    limited = False
    if max_paths <= 0:
        return rows, capped
    auth_adj, reach_adj = build_adjacency(codes, n)
    for i in range(n):
        if len(rows) >= max_paths:
            limited = True
            break
        if not auth_adj[i] and not reach_adj[i]:
            continue
        base = i * n
        # Pass A: nur auth_ok-Kanten
        prev, reason = _bfs(auth_adj, i, n, max_hops, FRONTIER_CAP)
        if reason == "frontier":
            capped = True
            continue  # zu dicht -> Subnetz-Fallback deckt ab
        for j in range(n):
            if len(rows) >= max_paths:
                limited = True
                break
            if i == j:
                continue
            c = codes[base + j]
            # Traversierbare Direkt-Verbindung? Dann braucht es keinen Pfad.
            if c == CODE_AUTH or c in CODE_REACH:
                continue
            if prev[j] != -1:
                p = _path_to(prev, i, j)
                rows.append({
                    "key": f"{ep_label(eps[i]['ip'], str(eps[i]['port']))}|"
                           f"{ep_label(eps[j]['ip'], str(eps[j]['port']))}",
                    "von": ep_label(eps[i]["ip"], str(eps[i]["port"])),
                    "nach": ep_label(eps[j]["ip"], str(eps[j]["port"])),
                    "pfad": " -> ".join(
                        ep_label(eps[k]["ip"], str(eps[k]["port"])) for k in p),
                    "art": "verifiziert (auth_ok)",
                })
        if len(rows) >= max_paths:
            limited = True
            break
        # Pass B: auth_ok + reachable (nur Ziele, die Pass A nicht fand)
        prev2, reason2 = _bfs(reach_adj, i, n, max_hops, FRONTIER_CAP)
        if reason2 == "frontier":
            capped = True
            continue
        for j in range(n):
            if len(rows) >= max_paths:
                limited = True
                break
            if i == j:
                continue
            c = codes[base + j]
            if c == CODE_AUTH or c in CODE_REACH:
                continue
            if prev[j] != -1 or prev2[j] == -1:
                continue
            p = _path_to(prev2, i, j)
            rows.append({
                "key": f"{ep_label(eps[i]['ip'], str(eps[i]['port']))}|"
                       f"{ep_label(eps[j]['ip'], str(eps[j]['port']))}",
                "von": ep_label(eps[i]["ip"], str(eps[i]["port"])),
                "nach": ep_label(eps[j]["ip"], str(eps[j]["port"])),
                "pfad": " -> ".join(
                    ep_label(eps[k]["ip"], str(eps[k]["port"])) for k in p),
                "art": "nur erreichbar",
            })
    if capped:
        log.warning("Pfadfindung: Frontier-Cap erreicht (dichte Quellen) - "
                    "Subnetz-Pfade decken die Reste ab")
    if limited:
        log.warning("Pfadfindung: max_paths=%d erreicht, weitere IP-Pfade "
                    "nicht gespeichert (Subnetz-Pfade decken die Reste ab)",
                    max_paths)
    return rows, capped


def find_net_paths(codes, eps, n, max_hops, log) -> list:
    """Netz-Graph: Knoten = Netze, Kante falls >=1 traversierbare IP-Kante.
    Pfade fuer ALLE Netzpärchen (auch direkt verbundene - die IP-Suche
    zeigt dann 'Netze direkt verbunden'). Return rows."""
    net_id = {}
    nets = []
    for e in eps:
        if e["net"] not in net_id:
            net_id[e["net"]] = len(nets)
            nets.append(e["net"])
    m = len(nets)
    if m == 0:
        return []
    auth_net = [set() for _ in range(m)]
    reach_net = [set() for _ in range(m)]
    for i in range(n):
        ni = net_id[eps[i]["net"]]
        base = i * n
        for j in range(n):
            if i == j:
                continue
            c = codes[base + j]
            if not c:
                continue
            nj = net_id[eps[j]["net"]]
            if ni == nj:
                continue
            if c == CODE_AUTH:
                auth_net[ni].add(nj)
                reach_net[ni].add(nj)
            elif c in CODE_REACH:
                reach_net[ni].add(nj)
    auth_adj = [sorted(s) for s in auth_net]
    reach_adj = [sorted(s) for s in reach_net]
    rows = []
    for ni in range(m):
        if not auth_adj[ni] and not reach_adj[ni]:
            continue
        direct = auth_net[ni] | reach_net[ni]
        # Direkt verbundene Netzpärchen: die IP-Suche zeigt dann
        # 'Netze direkt verbunden' statt 'kein Pfad'.
        for nj in direct:
            art = ("verifiziert (auth_ok)" if nj in auth_net[ni]
                   else "nur erreichbar")
            rows.append({
                "key": f"{nets[ni]}|{nets[nj]}",
                "von": nets[ni],
                "nach": nets[nj],
                "pfad": f"{nets[ni]} -> {nets[nj]}",
                "art": f"Netze direkt verbunden ({art})",
            })
        prev, _ = _bfs(auth_adj, ni, m, max_hops, 10 ** 9)
        for nj in range(m):
            if ni == nj or nj in direct:
                continue
            if prev[nj] != -1:
                p = _path_to(prev, ni, nj)
                rows.append({
                    "key": f"{nets[ni]}|{nets[nj]}",
                    "von": nets[ni],
                    "nach": nets[nj],
                    "pfad": " -> ".join(nets[k] for k in p),
                    "art": "verifiziert (auth_ok)",
                })
        prev2, _ = _bfs(reach_adj, ni, m, max_hops, 10 ** 9)
        for nj in range(m):
            if ni == nj or nj in direct:
                continue
            if prev[nj] != -1 or prev2[nj] == -1:
                continue
            p = _path_to(prev2, ni, nj)
            rows.append({
                "key": f"{nets[ni]}|{nets[nj]}",
                "von": nets[ni],
                "nach": nets[nj],
                "pfad": " -> ".join(nets[k] for k in p),
                "art": "nur erreichbar",
            })
    return rows


def write_paths_csv(path: str, rows: list) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["key", "von", "nach", "pfad", "art"])
        w.writeheader()
        w.writerows(rows)


def fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def write_matrix_csv(matrix_path: str, eps: list, codes: bytearray) -> None:
    n = len(eps)
    labels = [ep_label(e["ip"], str(e["port"])) for e in eps]
    with open(matrix_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["quelle\\ziel"] + labels)
        for i, s in enumerate(eps):
            row = [ep_label(s["ip"], str(s["port"]))]
            base = i * n
            for j in range(n):
                if i == j:
                    row.append("")
                    continue
                row.append(CODE_TO_SHORT.get(codes[base + j], ""))
            w.writerow(row)


def subnet_cell(agg_entry: dict) -> tuple:
    """Return (text, bg_color, font_color) fuer eine Netz-Matrix-Zelle."""
    if not agg_entry or agg_entry["tested"] == 0:
        text = "–" if agg_entry and agg_entry["skipped"] > 0 else ""
        return text, "E7E6E6", "595959"
    ok = agg_entry["ok"]
    tested = agg_entry["tested"]
    text = f"{ok}/{tested}"
    if ok > 0:
        return text, "C6EFCE", "006100"
    if agg_entry["port_only"] > 0:
        return text, "FFEB9C", "9C6500"
    return text, "FFC7CE", "9C0006"


def write_netz_matrix_csv(path: str, agg: dict, nets: list) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["quelle\\ziel"] + nets)
        for s in nets:
            row = [s]
            for t in nets:
                if s == t:
                    row.append("")
                    continue
                a = agg.get((s, t))
                if not a or a["tested"] == 0:
                    row.append("–" if a and a["skipped"] > 0 else "")
                else:
                    row.append(f"{a['ok']}/{a['tested']}")
            w.writerow(row)


def build_xlsx(xlsx_path: str, detail_header, detail_rows: list,
               detail_truncated: int, eps: list, codes: bytearray,
               labels: list, agg: dict = None, nets: list = None,
               src_ok=None, tgt_reach=None, tgt_auth=None,
               skipped_by_n24=None, matrix_limit: int = 0,
               ip_paths=None, net_paths=None) -> None:
    wb = Workbook()
    ws_search = wb.active
    ws_search.title = "Suche"
    ws_matrix = wb.create_sheet("Matrix")
    ws_netz = wb.create_sheet("Netz-Matrix") if agg else None
    ws_detail = wb.create_sheet("Detail")
    ws_subnets = wb.create_sheet("Subnetze")
    ws_quelle = wb.create_sheet("Quelle")
    ws_liste = wb.create_sheet("Liste")

    n = len(eps)
    last_col = get_column_letter(1 + n)
    last_row = 1 + n

    # ---- Liste (Dropdown-Quelle) --------------------------------------
    ws_liste["A1"] = "Endpunkt (Auswahl fuer Suche)"
    ws_liste["A1"].font = Font(bold=True)
    for i, lab in enumerate(labels, start=2):
        ws_liste.cell(row=i, column=1, value=lab)

    # ---- Suche ---------------------------------------------------------
    ws_search["A1"] = "SSH-Matrix - Suche VON -> NACH"
    ws_search["A1"].font = Font(bold=True, size=14)
    ws_search["A3"] = "VON"
    ws_search["A3"].font = Font(bold=True)
    ws_search["A4"] = "NACH"
    ws_search["A4"].font = Font(bold=True)
    ws_search["A5"] = "Ergebnis"
    ws_search["A5"].font = Font(bold=True)
    ws_search["C3"] = "Quell-Endpunkt (Dropdown / Tippen)"
    ws_search["C4"] = "Ziel-Endpunkt (Dropdown / Tippen)"
    ws_search["C5"] = "Status aus der Matrix"

    formula = (
        '=IF(OR($B$3="",$B$4=""),"",IFERROR('
        f'INDEX(Matrix!$B$2:${last_col}${last_row},'
        f'MATCH($B$3,Matrix!$A$2:$A${last_row},0),'
        f'MATCH($B$4,Matrix!$B$1:${last_col}$1,0)),"nicht getestet"))'
    )
    ws_search["B5"] = formula
    ws_search["B5"].font = Font(bold=True, size=12)

    dv = DataValidation(type="list", formula1=f"=Liste!$A$2:$A${last_row}",
                        allow_blank=True, showErrorMessage=False)
    ws_search.add_data_validation(dv)
    dv.add("B3")
    dv.add("B4")

    # Legende
    ws_search["A7"] = "Legende:"
    ws_search["A7"].font = Font(bold=True)
    legend = [
        ("OK", "Voller SSH-Login von A nach B erfolgreich"),
        ("AUTH", "SSH-Port offen, aber Login abgelehnt"),
        ("PORT", "Nur Port 22 erreichbar (kein SSH-Login getestet)"),
        ("CLOSED", "Port zu (Connection refused)"),
        ("UNREACH", "Netzwerk nicht erreichbar (Timeout/No Route)"),
        ("SRCERR", "Quelle war vom Kali aus nicht erreichbar"),
        ("NOTOOL", "Auf der Quelle kein ssh/nc/bash verfuegbar"),
        ("ERR", "Fehler (Detail-Sheet/error-Spalte beachten)"),
        ("SKIP", "Uebersprungen (Subnetz-Quota erreicht)"),
        ("?", "Undefiniert / nicht getestet"),
    ]
    for i, (code, desc) in enumerate(legend, start=8):
        ws_search.cell(row=i, column=1, value=code)
        ws_search.cell(row=i, column=2, value=desc)
        if code in STYLE:
            bg, fg = STYLE[code]
            ws_search.cell(row=i, column=1).fill = fill(bg)
            ws_search.cell(row=i, column=1).font = Font(color=fg, bold=True)

    ws_search.column_dimensions["A"].width = 12
    ws_search.column_dimensions["B"].width = 22
    ws_search.column_dimensions["C"].width = 40

    # ---- Pfad-Vorschlag (Mehrfach-Hop) ----------------------------------
    # Direkt-Ergebnis (B5) steht bereits; darunter: IP-Pfad, sonst
    # Netz-Pfad (Netz von VON/NACH via Quelle-Sheet), sonst "kein Pfad".
    ws_search["A6"] = "Pfad-Vorschlag"
    ws_search["A6"].font = Font(bold=True)
    net_von = 'INDEX(Quelle!$F:$F,MATCH($B$3,Quelle!$A:$A,0))'
    net_nach = 'INDEX(Quelle!$F:$F,MATCH($B$4,Quelle!$A:$A,0))'
    path_formula = (
        '=IFERROR(VLOOKUP($B$3&"|"&$B$4,Pfade!$A:$E,4,0),'
        f'IFERROR(VLOOKUP({net_von}&"|"&{net_nach},'
        "'Netz-Pfade'!$A:$D,3,0),\"kein Pfad\"))"
    )
    kind_formula = (
        '=IFERROR(VLOOKUP($B$3&"|"&$B$4,Pfade!$A:$E,5,0),'
        f'IFERROR(VLOOKUP({net_von}&"|"&{net_nach},'
        "'Netz-Pfade'!$A:$D,4,0),\"\"))"
    )
    ws_search["B6"] = path_formula
    ws_search["B6"].font = Font(bold=True)
    ws_search["C6"] = kind_formula
    ws_search["D6"] = "(Pfade: verifiziert = auth_ok, nur erreichbar = auth_fail/port_open)"
    ws_search["D6"].font = Font(italic=True, size=8, color="808080")

    # Bedingte Formatierung auf B6/C6: gruen=verifiziert, gelb=nur
    # erreichbar, grau=kein Pfad.
    from openpyxl.formatting.rule import FormulaRule
    for rng in ("B6:C6",):
        ws_search.conditional_formatting.add(
            rng, FormulaRule(formula=['ISNUMBER(SEARCH("verifiziert",$C6))'],
                             fill=fill("C6EFCE"), font=Font(color="006100")))
        ws_search.conditional_formatting.add(
            rng, FormulaRule(formula=['ISNUMBER(SEARCH("nur erreichbar",$C6))'],
                             fill=fill("FFEB9C"), font=Font(color="9C6500")))
        ws_search.conditional_formatting.add(
            rng, FormulaRule(formula=['$B6="kein Pfad"'],
                             fill=fill("D9D9D9"), font=Font(color="404040")))

    # ---- Matrix ---------------------------------------------------------
    n = len(eps)
    if matrix_limit and n > matrix_limit:
        # Host-Matrix bei sehr vielen Endpunkten ueberspringen (Excel-
        # Overkill: n^2 gestylte Zellen). matrix.csv + Netz-Matrix bleiben.
        ws_matrix["A1"] = (f"Host-Matrix uebersprungen: {n} Endpunkte > "
                           f"Limit {matrix_limit} - siehe matrix.csv "
                           f"und Sheet 'Netz-Matrix'.")
        ws_matrix["A1"].font = Font(bold=True, color="FF0000")
        ws_matrix.column_dimensions["A"].width = 60
    else:
        ws_matrix["A1"] = "Quelle \\ Ziel"
        ws_matrix["A1"].font = Font(bold=True)
        for j, lab in enumerate(labels, start=2):
            c = ws_matrix.cell(row=1, column=j, value=lab)
            c.font = Font(bold=True, size=8)
            c.alignment = Alignment(textRotation=90, horizontal="center")
            ws_matrix.column_dimensions[get_column_letter(j)].width = 7
        for i, s in enumerate(eps):
            ws_matrix.cell(row=i + 2, column=1,
                           value=ep_label(s["ip"], str(s["port"]))).font = Font(bold=True)
            base = i * n
            for j in range(n):
                if i == j:
                    continue
                code = CODE_TO_SHORT.get(codes[base + j], "?")
                cell = ws_matrix.cell(row=i + 2, column=j + 2, value=code)
                cell.alignment = Alignment(horizontal="center")
                if code in STYLE:
                    bg, fg = STYLE[code]
                    cell.fill = fill(bg)
                    cell.font = Font(color=fg, size=8)
        ws_matrix.column_dimensions["A"].width = 15
        ws_matrix.freeze_panes = "B2"

    # ---- Netz-Matrix ---------------------------------------------------
    if ws_netz is not None and nets:
        ws_netz["A1"] = "Quell-Netz \\ Ziel-Netz"
        ws_netz["A1"].font = Font(bold=True)
        for j, t_net in enumerate(nets, start=2):
            c = ws_netz.cell(row=1, column=j, value=t_net)
            c.font = Font(bold=True, size=8)
            c.alignment = Alignment(textRotation=90, horizontal="center")
            ws_netz.column_dimensions[get_column_letter(j)].width = 11
        for i, s_net in enumerate(nets, start=2):
            ws_netz.cell(row=i, column=1, value=s_net).font = Font(bold=True, size=8)
            for j, t_net in enumerate(nets, start=2):
                if s_net == t_net:
                    continue
                text, bg, fg = subnet_cell(agg.get((s_net, t_net)))
                cell = ws_netz.cell(row=i, column=j, value=text)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = fill(bg)
                cell.font = Font(color=fg, size=8)
        ws_netz.column_dimensions["A"].width = 18
        ws_netz.freeze_panes = "B2"

    # ---- Detail (gekappt auf detail_max; volle Daten in detail.csv) ----
    if detail_header:
        ws_detail.append(detail_header)
        for cell in ws_detail[1]:
            cell.font = Font(bold=True)
        for r in detail_rows:
            ws_detail.append([r.get(h, "") for h in detail_header])
        n_rows = len(detail_rows) + 1
        if detail_truncated:
            note = [f"... weitere {fmt_num(detail_truncated)} Zeilen in "
                    "detail.csv (Detail-Sheet gekappt)"]
            note += [""] * (len(detail_header) - 1)
            ws_detail.append(note)
            n_rows += 1
        ws_detail.auto_filter.ref = f"A1:{get_column_letter(len(detail_header))}{n_rows}"
        ws_detail.freeze_panes = "A2"

        status_col = detail_header.index("status") + 1
        col_letter = get_column_letter(status_col)
        rng = f"{col_letter}2:{col_letter}{len(detail_rows) + 1}"
        for st, code in CODE.items():
            if code not in STYLE:
                continue
            bg, fg = STYLE[code]
            ws_detail.conditional_formatting.add(
                rng,
                CellIsRule(operator="equal", formula=[f'"{st}"'],
                           fill=fill(bg), font=Font(color=fg)))
        widths = {"timestamp": 22, "source_ip": 15, "source_port": 9, "source_label": 18,
                  "src_24": 13, "src_16": 13, "src_net": 16, "target_ip": 15,
                  "target_port": 9, "target_label": 18, "tgt_24": 13, "tgt_16": 13,
                  "tgt_net": 16, "direction": 9, "method": 10, "status": 13,
                  "latency_ms": 11, "error": 60}
        for idx, h in enumerate(detail_header, start=1):
            ws_detail.column_dimensions[get_column_letter(idx)].width = widths.get(h, 14)

    # ---- Subnetze -------------------------------------------------------
    src_ok = src_ok or set()
    tgt_reach = tgt_reach or set()
    tgt_auth = tgt_auth or set()
    skipped_by_n24 = skipped_by_n24 or Counter()

    ep_by_n24 = defaultdict(list)
    for e in eps:
        ep_by_n24[e["n24"]].append(e)

    ws_subnets.append(["/24-Subnetz", "Endpunkte", "Quelle mit Login-Erfolg",
                       "als Ziel erreichbar (Login oder Port)", "als Ziel mit Login-Erfolg",
                       "skipped (Quota)"])
    for cell in ws_subnets[1]:
        cell.font = Font(bold=True)
    for n24 in sorted(ep_by_n24):
        members = ep_by_n24[n24]
        keys = [(m["ip"], m["port"]) for m in members]
        row = [
            n24,
            len(members),
            sum(1 for k in keys if k in src_ok),
            sum(1 for k in keys if k in tgt_reach),
            sum(1 for k in keys if k in tgt_auth),
            skipped_by_n24.get(n24, 0),
        ]
        ws_subnets.append(row)
    ws_subnets.auto_filter.ref = f"A1:F{len(ep_by_n24) + 1}"
    ws_subnets.freeze_panes = "A2"
    for col, w in zip("ABCDEF", (15, 10, 28, 34, 26, 14)):
        ws_subnets.column_dimensions[col].width = w

    # ---- Quelle ---------------------------------------------------------
    # Spalte A = Endpunkt-Label (ep_label: IP oder IP:PORT) - die
    # Suche-Formel matcht B3/B4 (Dropdown-Werte) gegen diese Spalte.
    ws_quelle.append(["Endpunkt", "Port", "Label", "/24", "/16", "Netz"])
    for cell in ws_quelle[1]:
        cell.font = Font(bold=True)
    for e in eps:
        ws_quelle.append([ep_label(e["ip"], str(e["port"])), e["port"],
                          e["label"], e["n24"], e["n16"], e.get("net", "")])
    ws_quelle.auto_filter.ref = f"A1:F{len(eps) + 1}"
    ws_quelle.freeze_panes = "A2"
    for col, w in zip("ABCDEF", (16, 8, 24, 13, 13, 16)):
        ws_quelle.column_dimensions[col].width = w

    # ---- Pfade (IP-Pfade) ----------------------------------------------
    ws_pfade = wb.create_sheet("Pfade")
    for col, h in enumerate(("Schluessel", "VON", "NACH", "Pfad", "Art"), start=1):
        ws_pfade.cell(row=1, column=col, value=h)
        ws_pfade.cell(row=1, column=col).font = Font(bold=True)
        ws_pfade.column_dimensions[get_column_letter(col)].width = (
            34, 18, 18, 60, 24)[col - 1]
    if ip_paths:
        for r in ip_paths:
            ws_pfade.append([r["key"], r["von"], r["nach"], r["pfad"], r["art"]])
        ws_pfade.auto_filter.ref = f"A1:E{len(ip_paths) + 1}"
    ws_pfade.freeze_panes = "A2"

    # ---- Netz-Pfade (Subnetz-Fallback) ---------------------------------
    ws_netz_pfade = wb.create_sheet("Netz-Pfade")
    for col, h in enumerate(("Schluessel", "VON-Netz", "NACH-Netz", "Pfad", "Art"),
                            start=1):
        ws_netz_pfade.cell(row=1, column=col, value=h)
        ws_netz_pfade.cell(row=1, column=col).font = Font(bold=True)
        ws_netz_pfade.column_dimensions[get_column_letter(col)].width = (
            36, 20, 20, 70, 24)[col - 1]
    if net_paths:
        for r in net_paths:
            ws_netz_pfade.append([r["key"], r["von"], r["nach"], r["pfad"], r["art"]])
        ws_netz_pfade.auto_filter.ref = f"A1:E{len(net_paths) + 1}"
    ws_netz_pfade.freeze_panes = "A2"

    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser(description="Erzeugt Matrix-CSV + report.xlsx aus detail.csv")
    ap.add_argument("--version", action="version",
                    version=f"%(prog)s {VERSION} (entwickelt von {AUTHOR})")
    ap.add_argument("--detail", required=True, help="Pfad zu detail.csv")
    ap.add_argument("--out", default="ssh_matrix_out",
                    help="Ausgabe-Verzeichnis (Default: ssh_matrix_out)")
    ap.add_argument("--matrix-name", default="matrix.csv")
    ap.add_argument("--xlsx-name", default="report.xlsx")
    ap.add_argument("--detail-max", type=int, default=50000,
                    help="Detail-Sheet auf max N Zeilen kappen "
                         "(Default: 50000, Excel-Limit 1048576; 0 = alle)")
    ap.add_argument("--matrix-limit", type=int, default=2000,
                    help="Host-Matrix-Sheet ab N Endpunkten ueberspringen "
                         "(Default: 2000; 0 = nie)")
    ap.add_argument("--paths-hops", type=int, default=6,
                    help="Max. Hops fuer Pfad-Vorschlaege (Default: 6)")
    ap.add_argument("--paths-max", type=int, default=300000,
                    help="Max. gespeicherte IP-Pfade (Default: 300000; 0 = aus)")
    args = ap.parse_args()

    print_banner()

    if not os.path.exists(args.detail):
        print(f"FEHLER: {args.detail} existiert nicht", file=sys.stderr)
        sys.exit(1)

    detail_max = args.detail_max if args.detail_max > 0 else 1048576

    # Streaming: detail.csv wird NIE komplett in den RAM geladen.
    analysis = analyze_detail(args.detail)
    if analysis["rows_total"] == 0:
        print(f"FEHLER: {args.detail} enthaelt keine Daten", file=sys.stderr)
        sys.exit(1)

    eps = analysis["eps"]
    labels = [ep_label(e["ip"], str(e["port"])) for e in eps]
    n = len(eps)

    codes, detail_header, detail_rows, detail_truncated = pass2_codes(
        args.detail, analysis["id_of"], n, detail_max)

    os.makedirs(args.out, exist_ok=True)
    matrix_path = os.path.join(args.out, args.matrix_name)
    netz_matrix_path = os.path.join(args.out, "netz_matrix.csv")
    pfade_path = os.path.join(args.out, "pfade.csv")
    netz_pfade_path = os.path.join(args.out, "netz_pfade.csv")
    xlsx_path = os.path.join(args.out, args.xlsx_name)

    # Pfadfindung: IP-Pfade (verifiziert zuerst, sonst erreichbar) +
    # Subnetz-Fallback. Nur bei Bedarf (paths_max > 0).
    ip_paths = []
    net_paths = []
    if args.paths_max > 0 and args.paths_hops > 0:
        ip_paths, capped = find_ip_paths(codes, eps, n, args.paths_hops,
                                         args.paths_max, logging.getLogger("ssh_matrix_report"))
        net_paths = find_net_paths(codes, eps, n, args.paths_hops,
                                   logging.getLogger("ssh_matrix_report"))
        write_paths_csv(pfade_path, ip_paths)
        write_paths_csv(netz_pfade_path, net_paths)

    write_matrix_csv(matrix_path, eps, codes)
    write_netz_matrix_csv(netz_matrix_path, analysis["agg"], analysis["nets"])
    build_xlsx(xlsx_path, detail_header, detail_rows, detail_truncated,
               eps, codes, labels, analysis["agg"], analysis["nets"],
               analysis["src_ok"], analysis["tgt_reach"], analysis["tgt_auth"],
               analysis["skipped_by_n24"], args.matrix_limit,
               ip_paths, net_paths)

    print(f"Matrix-CSV    : {matrix_path}")
    print(f"Netz-Matrix-CSV: {netz_matrix_path}")
    print(f"Excel-Report  : {xlsx_path}")
    print(f"  Endpunkte : {n}")
    print(f"  Netze     : {len(analysis['nets'])}")
    print(f"  Zeilen    : {analysis['rows_total']}")
    if args.paths_max > 0 and args.paths_hops > 0:
        print(f"  IP-Pfade  : {len(ip_paths)}")
        print(f"  Netz-Pfade: {len(net_paths)}")
        if not ip_paths and not net_paths:
            print("  Hinweis   : keine Mehrfach-Hop-Pfade gefunden "
                  "(alle Paare direkt oder unerreichbar)")
    if detail_truncated:
        print(f"  Hinweis   : Detail-Sheet auf {len(detail_rows)} Zeilen "
              f"gekappt ({fmt_num(detail_truncated)} weitere in detail.csv)")


if __name__ == "__main__":
    main()
